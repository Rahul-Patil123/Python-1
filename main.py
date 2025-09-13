import openpyxl

inv_list = openpyxl.load_workbook("inventory.xlsx")
product_details = inv_list["Sheet1"]

product_per_company = {}
inventory_per_company = {}
inventory_less_10 = {}

for product_row in range(2, product_details.max_row + 1):
    supplier_name = product_details.cell(product_row, 4).value
    supplier_name = product_details.cell(product_row, 4).value
    inventory_value = product_details.cell(product_row, 2).value
    inventory_price = product_details.cell(product_row, 3).value
    price_per_inv = product_details.cell(product_row, 5)
    product_id = int(product_details.cell(product_row, 1).value)
    inventory = int(product_details.cell(product_row, 2).value)
    
    if supplier_name in product_per_company:
        product_per_company[supplier_name] += 1
    else:
        product_per_company[supplier_name] = 1
        
    if supplier_name in inventory_per_company:
        inventory_per_company[supplier_name] += inventory_value * inventory_price
    else:
        inventory_per_company[supplier_name] = inventory_value * inventory_price
        
    if inventory < 10:
        inventory_less_10[product_id] = inventory
    
    price_per_inv.value = inventory_value * inventory_price

print("These are total products per company: ", product_per_company)
print("These are total inventory per company: ", inventory_per_company)
print("These are product whose inventory less than 10: ", inventory_less_10)
inv_list.save("inventory_solution.xlsx")